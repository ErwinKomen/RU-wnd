"""
Correct the Excel WGD file

This version created by Erwin R. Komen
Date: 29/oct/2020

Example:
    python diacorrect.py 
        -i "D:/Data Files/TG/Dialecten/Data/2020/WGD/Roeland/WGD-Rivierengebied-mens-20201028_EK.xlsx" 
        -o "D:/Data Files/TG/Dialecten/Data/2020/WGD/Roeland/WGD-Rivierengebied-mens-20201029_EK.xlsx"

"""

import sys, getopt, os.path, importlib
import io, sys, os
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook
from io import StringIO

# My own stuff
from utils import ErrHandle

# Make available error handling
errHandle = ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 29/oct/2020    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
    flInput = ''        # input file name: XML with author definitions
    flOutput = ''       # output file name
    readonly = True

    try:
        sSyntax = prgName + ' -i <Excel WGD input file> -o <output file> [-w]'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hi:o:w", ["-ifile=", "-ofile"])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--ifile"):
                flInput = arg
            elif opt in ("-o", "--ofile"):
                flOutput = arg
            elif opt in ("-w", "--write"):
                readonly = False
        # Check if all arguments are there
        if (flInput == '' or flOutput == ''):
            errHandle.DoError(sSyntax)

        # Continue with the program
        errHandle.Status('Input is "' + flInput + '"')
        errHandle.Status('Output is "' + flOutput + '"')

        # Call the function that does the job
        oArgs = dict(input=flInput, output=flOutput, readonly=readonly)
        if (not do_dialect_correct(oArgs)) :
            errHandle.DoError("Could not complete")
            return False
        
            # All went fine    
        errHandle.Status("Ready")
    except:
        # act
        errHandle.DoError("main")
        return False

# ----------------------------------------------------------------------------------
# Name :    do_dialect_correct
# Goal :    Perfrom dialect corrections
# History:
# 29/oct/2020    ERK Created
# ----------------------------------------------------------------------------------
def do_dialect_correct(oArgs):
    """Perfrom dialect correction"""

    # Defaults
    flInput = ""
    flOutput = ""
    lAuthor = []
    read_only = True

    count = 0

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]
        if "readonly" in oArgs: read_only = oArgs["readonly"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Read the input file as excel
        wb = openpyxl.load_workbook(flInput, read_only=read_only)
        # Get the main worksheet
        ws = wb.active

        # Iterate through the rows
        bFirst = True
        header = {}

        for row in ws.iter_rows(min_row=1, min_col=1):
            if bFirst:
                # Expect header: interpret it
                for idx, cell in enumerate(row):
                    if cell.value != None:
                        sValue = cell.value.strip("\t").lower()
                        header[sValue] = idx
                # Expecting and skipping header
                bFirst = False
            elif row[0].value != None:
                # We are in a row in which there is some information
                rownumber = row[0].row
                if rownumber % 1000 == 0:
                    errHandle.Status("Row: {}".format(rownumber))

                if rownumber >= 293:
                    iStop = 1
                
                # (1) Correct 'lijstnummer'
                plaats = row[header['plaats-corr']].value
                lijstnummer = row[header['lijstnummer']].value
                if plaats == "Bruchem":
                    if lijstnummer == 9 or lijstnummer == 10:
                        # Adjust the 'lijstnummer'
                        lijstnummer += 10
                        if not read_only:
                            row[header['lijstnummer']].value = lijstnummer

                # (2) Deal with 'volkskundig'
                lemmatitel = row[header['lemmatitel']].value
                if lemmatitel != None and 'volkskundig' in lemmatitel:
                    if not read_only:
                        # Adapt the lemmatitel itself
                        row[header['lemmatitel']].value = lemmatitel.replace("volkskundig", "").strip()
                        # Add a note in 'opmerkingen'
                        row[header['opmerkingen']].value = "volkskundig"

                # (3) Deal with (n)
                standaardspelling = row[header['standaardspelling']].value
                dialectwoord = row[header['dialectwoord']].value
                if standaardspelling != None and dialectwoord != None:
                    if not read_only:
                        if "(n)" in standaardspelling and not "(n)" in dialectwoord:
                            if "n" in dialectwoord:
                                # ete  - ete(n) ==> ete
                                row[header['opmerkingen']].value = standaardspelling
                                row[header['standaardspelling']].value = standaardspelling.replace("(n)", "n")
                            else:
                                # eten - ete(n) ==> eten
                                # Remove the (n)
                                row[header['opmerkingen']].value = standaardspelling
                                row[header['standaardspelling']].value = standaardspelling.replace("(n)", "")
                    #if "(n)" in dialectwoord and not "(n)" in standaardspelling:
                    #    if not read_only:
                    #        row[header['standaardspelling']].value = "{}(n)".format(standaardspelling)
                    #elif "n" in dialectwoord and "(n)" in standaardspelling:
                    #    if not read_only:
                    #        row[header['standaardspelling']].value = standaardspelling.replace("(n)", "n")
                    #elif not "(n)" in dialectwoord and "(n)" in standaardspelling:
                    #    if not read_only:
                    #        row[header['standaardspelling']].value = standaardspelling.replace("(n)", "")

        # Save the workbook differently
        wb.save(flOutput)
        # Close the workbook
        wb.close()


        # Return positively
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("do_dialect_correct")
        return False




# ----------------------------------------------------------------------------------
# Goal :    If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
    # Call the main function with two arguments: program name + remainder
    main(sys.argv[0], sys.argv[1:])
