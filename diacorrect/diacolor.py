"""
Color empty cells in the Excel WGD file

This version created by Erwin R. Komen
Date: 2/dec/2021

Example:
    python diacolor.py -w
        -i "D:/Data Files/TG/Dialecten/Data/2021/WGD/erwin/WGD-Veluwe-mens-Erwin.xlsx" 
        -o "D:/Data Files/TG/Dialecten/Data/2021/WGD/erwin/WGD-Veluwe-mens-Erwin-out.xlsx"

"""

import sys, getopt, os.path, importlib
import io, sys, os
import openpyxl
import copy
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from io import StringIO

# My own stuff
from utils import ErrHandle

# Make available error handling
errHandle = ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 2/dec/2021    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
    flInput = ''        # input file name: XML with author definitions
    flOutput = ''       # output file name
    readonly = True

    try:
        sSyntax = prgName + ' -i <Excel WGD input file> -o <output file> [-w]'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hi:o:w", ["-ifile=", "-ofile"])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--ifile"):
                flInput = arg
            elif opt in ("-o", "--ofile"):
                flOutput = arg
            elif opt in ("-w", "--write"):
                readonly = False
        # Check if all arguments are there
        if (flInput == '' or flOutput == ''):
            errHandle.DoError(sSyntax)

        # Continue with the program
        errHandle.Status('Input is "' + flInput + '"')
        errHandle.Status('Output is "' + flOutput + '"')

        # Call the function that does the job
        oArgs = dict(input=flInput, output=flOutput, readonly=readonly)
        if (not do_dialect_color(oArgs)) :
            errHandle.DoError("Could not complete")
            return False
        
            # All went fine    
        errHandle.Status("Ready")
    except:
        # act
        errHandle.DoError("main")
        return False

# ----------------------------------------------------------------------------------
# Name :    do_dialect_color
# Goal :    Perfrom coloring of empty cells
# History:
# 2/dec/2021    ERK Created
# ----------------------------------------------------------------------------------
def do_dialect_color(oArgs):
    """Perfrom dialect correction"""

    # Defaults
    flInput = ""
    flOutput = ""
    lAuthor = []
    read_only = True
    columns = 12
    count = 0
    column_dict = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    orangeFill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type='solid')

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]
        if "readonly" in oArgs: read_only = oArgs["readonly"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Read the input file as excel
        errHandle.Status("Reading the Excel...")
        wb = openpyxl.load_workbook(flInput, read_only=read_only)
        errHandle.Status("Continuing...")
        # Get the main worksheet
        ws = wb.active

        # Iterate through the rows
        bFirst = True
        header = {}
        lst_msg = []

        for row in ws.iter_rows(min_row=1, min_col=1):
            if bFirst:
                # Expect header: interpret it
                for idx, cell in enumerate(row):
                    if cell.value != None:
                        sValue = cell.value.strip("\t").lower()
                        header[sValue] = idx
                # Expecting and skipping header
                bFirst = False
                errHandle.Status("Dealt with first line...")
            elif row[0].value != None:
                # We are in a row in which there is some information
                rownumber = row[0].row
                if rownumber % 1000 == 0:
                    errHandle.Status("Row: {}".format(rownumber))

                if rownumber >= 293:
                    iStop = 1

                # Look for empty cells in this row
                bHasEmpty = False
                row_msg = []
                for idx in range(1,columns):
                    col = idx - 1
                    if row[col].value == None or row[col].value == "":
                        # Possibly start message
                        if len(row_msg) == 0:
                            row_msg.append("Empty cells in Row {}: ".format(rownumber))
                        # Always provide the column
                        row_msg.append(" {}".format(column_dict[col]))
                        # Color this cell
                        row[col].fill = orangeFill
                if len(row_msg) > 0:
                    sMsg = "".join(row_msg)
                    lst_msg.append(sMsg)
                
        # Save the message list
        file_msg = flInput.replace(".xlsx", "-LegeCellen.txt")
        msg = "\n".join(lst_msg)
        with open(file_msg, "w", encoding="utf-8") as fp:
            fp.write(msg)
 
        # Save the workbook differently
        wb.save(flOutput)
        # Close the workbook
        wb.close()

        # Return positively
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("do_dialect_color")
        return False




# ----------------------------------------------------------------------------------
# Goal :    If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
    # Call the main function with two arguments: program name + remainder
    main(sys.argv[0], sys.argv[1:])
