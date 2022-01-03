"""
Definition of views.
"""

from django.contrib.auth import login, authenticate
from django.contrib.auth.models import User, Group
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView, View
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.template import RequestContext, loader
from django.template.loader import render_to_string
from django.db import connection
from django.db.models import Q
from django.db.models.functions import Lower
from django.http import JsonResponse
from datetime import datetime
from xml.dom import minidom
from operator import itemgetter
import xml.etree.ElementTree as ET
import os
import operator
import re
import fnmatch
import csv
import codecs
import copy
import sys
from wald.dictionary.models import *
from wald.dictionary.forms import *
from wald.mapview.views import MapView
from wald.settings import APP_PREFIX, WSGI_FILE
from wald.dictionary.conversion import rd_to_wgs, wgs_to_rd

# Global variables
paginateSize = 10
paginateLocations = 2
paginateEntries = 100
# OLD: paginateValues = (1000, 500, 250, 100, 50, 40, 30, 20, 10, )
paginateValues = (100, 50, 20, 10, 5, 2, 1, )
outputColumns = ['begrip', 'trefwoord', 'dialectopgave', 'Kloekecode', 'aflevering', 'bronnenlijst']
bDebug = True

THIS_DICTIONARY = "e-WALD"

# User group names important for anything
app_userplus = "ewald_userplus" 
app_editor = "ewald_editor"

# General help functions
def order_queryset_by_sort_order(get, qs, sOrder = 'gloss'):
    """Change the sort-order of the query set, depending on the form field [sortOrder]

    This function is used by EntryListView.
    The value of [sortOrder] is 'woord' by default.
    [sortOrder] is a hidden field inside the "adminsearch" html form in the template admin_gloss_list.html
    Its value is changed by clicking the up/down buttons in the second row of the search result table
    """

    def get_string_from_tuple_list(lstTuples, number):
        """Get the string value corresponding to a number in a list of number-string tuples"""
        sBack = [tup[1] for tup in lstTuples if tup[0] == number]
        return sBack

    # Helper: order a queryset on field [sOrder], which is a number from a list of tuples named [sListName]
    def order_queryset_by_tuple_list(qs, sOrder, sListName):
        """Order a queryset on field [sOrder], which is a number from a list of tuples named [sListName]"""

        # Get a list of tuples for this sort-order
        tpList = build_choice_list(sListName)
        # Determine sort order: ascending is default
        bReversed = False
        if (sOrder[0:1] == '-'):
            # A starting '-' sign means: descending order
            sOrder = sOrder[1:]
            bReversed = True

        # Order the list of tuples alphabetically
        # (NOTE: they are alphabetical from 'build_choice_list()', except for the values 0,1)
        tpList = sorted(tpList, key=operator.itemgetter(1))
        # Order by the string-values in the tuple list
        return sorted(qs, key=lambda x: get_string_from_tuple_list(tpList, getattr(x, sOrder)), reverse=bReversed)

    # Set the default sort order
    # sOrder = 'gloss'  # Default sort order if nothing is specified
    # See if the form contains any sort-order information
    if ('sortOrder' in get and get['sortOrder'] != ''):
        # Take the user-indicated sort order
        sOrder = get['sortOrder']

    # The ordering method depends on the kind of field:
    # (1) text fields are ordered straightforwardly
    # (2) fields made from a choice_list need special treatment
    if (sOrder.endswith('handedness')):
        ordered = order_queryset_by_tuple_list(qs, sOrder, "Handedness")
    else:
        # Use straightforward ordering on field [sOrder]
        ordered = qs.order_by(Lower(sOrder))

    # return the ordered list
    return ordered

def get_item_list(lVar, lFun, qs):
    """Turn the queryset [qs] into a list of Items that have first and last information"""

    lItem = []
    oErr = ErrHandle()
    try:
        # Initialize the variables whose changes are important
        oVariable = {}
        for i, key in enumerate(lVar):
            oVariable[key] = "" # {'name': key, 'fun': lFun[i]}
        iLast = len(qs)-1
        # Iterate over the entries looking for first, last etc
        for i, entry in enumerate(qs):
            bIsLastEntry = (i==iLast)
            oItem = {'entry': entry}
            for k in lVar:
                oItem[k] = {'first':False, 'last':False}
            bIsDict = isinstance(entry, dict)
            bVarIsLast = False
            # Check for changes in all the variables
            for j, k in enumerate(lVar):
                fun = lFun[j]
                if callable(fun):
                    sValue = fun(entry)
                else:
                    for idx, val in enumerate(fun):
                        if idx==0:
                            if bIsDict:
                                sValue = entry[val]
                            else:
                                sValue = getattr(entry, val)
                        else:
                            if bIsDict:
                                sValue = sValue[val]
                            else:
                                sValue = getattr(sValue, val)
                # Check for changes in the value of the variable 
                # if sValue != oVariable[k]:
                if sValue != oVariable[k] or bVarIsLast or (i>0 and lItem[i-1][k]['last']):
                    # Check if the previous one's [last] must be changed
                    if oVariable[k] != "": lItem[i-1][k]['last'] = True
                    # Adapt the current one's [first] property
                    oItem[k]['first']= True
                    # Adapt the variable
                    oVariable[k] = sValue      
                    # Indicate that the next ones should be regarded as 'last'
                    bVarIsLast = True      
                # Check if this is the last
                if bIsLastEntry: oItem[k]['last'] = True
            # Add this object to the list of items
            lItem.append(oItem)
    except:
        oErr.DoError("get_item_list error")
        lItem = []

    # Return the list we have made
    return lItem

def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/index.html',
        {
            'title': THIS_DICTIONARY,
            'year':datetime.now().year,
            'is_app_userplus': user_is_ingroup(request, app_userplus),
            'is_app_editor': user_is_ingroup(request, app_editor),
        }
    )

def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/contact.html',
        {
            'title':'{} contact'.format(THIS_DICTIONARY),
            'message':'Henk van den Heuvel (H.vandenHeuvel@Let.ru.nl)',
            'year':datetime.now().year,
            'is_app_userplus': user_is_ingroup(request, app_userplus),
            'is_app_editor': user_is_ingroup(request, app_editor),
        }
    )

def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render( request, 'dictionary/about.html',
        {   'title':'{} informatie'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'year':datetime.now().year,
            'is_app_userplus': user_is_ingroup(request, app_userplus),
            'is_app_editor': user_is_ingroup(request, app_editor),
        }
    )

def guide(request):
    """Renders the 'guide' page."""
    assert isinstance(request, HttpRequest)
    return render( request, 'dictionary/guide.html',
        {   'title':'{} handleiding'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'dic_abbr': THIS_DICTIONARY,
            'year':datetime.now().year,
            'is_app_userplus': user_is_ingroup(request, app_userplus),
            'is_app_editor': user_is_ingroup(request, app_editor),
        }
    )

def login_as_user(request, user_id):
    assert isinstance(request, HttpRequest)

    # Find out who I am
    supername = request.user.username
    super = User.objects.filter(username__iexact=supername).first()
    if super == None:
        return nlogin(request)

    # Make sure that I am superuser
    if super.is_staff and super.is_superuser:
        user = User.objects.filter(username__iexact=user_id).first()
        if user != None:
            # Perform the login
            login(request, user)
            return HttpResponseRedirect(reverse("home"))

    return home(request)

def signup(request):
    """Provide basic sign up and validation of it """

    bStaffDefault = False

    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            # Save the form
            form.save()
            # Create the user
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')

            if bStaffDefault:
                # also make sure that the user gets into the STAFF,
                #      otherwise he/she may not see the admin pages
                user = authenticate(username=username, 
                                    password=raw_password,
                                    is_staff=True)
                user.is_staff = True
                user.save()
                # Add user to the "passim_user" group
                gQs = Group.objects.filter(name="wald_user")
                if gQs.count() > 0:
                    g = gQs[0]
                    g.user_set.add(user)
            else:
                user = authenticate(username=username, 
                                    password=raw_password,
                                    is_staff=False)
            # Log in as the user
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'dictionary/signup.html', {'form': form})

def user_is_ingroup(request, sGroup):
    # Is this user part of the indicated group?
    user = User.objects.filter(username=request.user.username).first()

    # Only look at group if the user is known
    if user == None:
        glist = []
    else:
        glist = [x.name for x in user.groups.all()]

        # Only needed for debugging
        if bDebug:
            ErrHandle().Status("User [{}] is in groups: {}".format(user, glist))
    # Evaluate the list
    bIsInGroup = (sGroup in glist)
    return bIsInGroup

def do_repair(request):
    """Renders the REPAIR page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'dictionary/repair.html',
        {   'title':'{} reparatie'.format(THIS_DICTIONARY),
            'message':'Radboud Universiteit Nijmegen - Dialectenwoordenboek.',
            'year':datetime.now().year,
        }
    )

def adapt_search(val):
    # First trim
    val = val.strip()    
    # Adapt for the use of '#'
    if '#' in val:
        # val = val.replace('#', '\\b.*')
        #val = r'(^|\b)' + fnmatch.translate(val.replace('#', '')) + r'($|\b)'
        val = r'(^|(.*\b))' + val.replace('#', r'((\b.*)|$)') # + r'((\b.*)|$)'
    else:
        val = '^' + fnmatch.translate(val) + '$'
    return val

def export_csv(qs, sFileName):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+sFileName+'.csv"'
     
    # Create a writer for the CSV
    # writer = csv.writer(response, csv.excel, delimiter='\t')
    writer = csv.writer(response, csv.excel_tab)
    # BOM to indicate that this is UTF8
    response.write(u'\ufeff'.encode('utf8'))
    # Output the first row with the headings -- these take 'entry' as basis
    fields = outputColumns
    # Output the first row with the headings
    writer.writerow(fields)
    # Walk through the queryset
    for obj in qs:
        writer.writerow(obj.get_row())

    return response

def export_xlsx(qs, sFileName):
    import openpyxl
    from openpyxl.utils.cell import get_column_letter
    from openpyxl import Workbook

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+sFileName+'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = sFileName

    row_num = 0

    columns = outputColumns
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num]
        # c.style.font.bold = True
        c.font = openpyxl.styles.Font(bold=True)
        # Set column width
        # ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
        # ws.column_dimensions[get_column_letter(col_num+1)].bestFit = True
        ws.column_dimensions[get_column_letter(col_num+1)].width = 20.0

    # Walk the queryset
    for obj in qs:
        row_num += 1
        row = obj.get_row()
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.alignment = openpyxl.styles.Alignment(wrap_text=True)

    wb.save(response)
    return response

def export_html(qs, sFileName):
    # Create a string with the HTML contents
    arOut = []
    arOut.append("<html><head><meta charset='utf-8' /></head><body><table><thead><tr>")
    fields = outputColumns
    for f in fields:
        arOut.append("<th>"+f+"</th>")
    arOut.append("</tr></thead>")
    arOut.append("<tbody>")
    # Walk the contents
    for obj in qs:
        sRow = "<tr><td>"+obj.lemma.gloss+"</td><td>"+obj.trefwoord.woord+"</td><td>"+obj.woord+"</td><td>"+obj.dialect.nieuw+"</td><td>"+obj.aflevering.naam+"</td><tr>"
        arOut.append(sRow)
    content = "\n".join(arOut)
    # Create the HttpResponse object with the appropriate header.
    response = HttpResponse(content, content_type='text/html')
    response['Content-Disposition'] = 'attachment; filename="'+sFileName+'.htm"'
     
    return response

def do_repair_start(request):
    """Start up the repair action"""
    sRepairType = request.GET.get('repairtype', '')

    # Formulate a response
    data = {'status': 'done'}

    # Remove any previous repair objects of this type
    Repair.objects.filter(repairtype=sRepairType).delete()
    # Retrieve the Repair object with the correct type
    oRepair = Repair(repairtype=sRepairType)
    oRepair.save()
    if sRepairType == "lemma":
        bResult = do_repair_lemma(oRepair)
        if not bResult:
            data['status'] = "error"
    elif sRepairType == "entrydescr":
        bResult = do_repair_entrydescr(oRepair)
        if not bResult:
            data['status'] = "error"
    elif sRepairType == "clean":
        bResult = do_repair_clean(oRepair)
        if not bResult:
            data['status'] = "error"

    # Return this response
    return JsonResponse(data)

def do_repair_progress(request):
    """Show the progress of where we are in repairing"""

    sRepairType = request.GET.get('repairtype', '')

    # Formulate a response
    data = {'status': 'not found'}

    try:
        # Get the repair object
        qs = Repair.objects.filter(repairtype=sRepairType)
        if qs != None and len(qs) > 0:
            oRepair = qs[0]
            data['status'] = oRepair.status
    except:
        data['status'] = 'skipping a beat (repair_progress)'

    # Return this response
    return JsonResponse(data)

def import_csv_start(request):
    # Formulate a response
    data = {'status': 'done'}
    oErr = ErrHandle()

    try:
        # x = request.POST
        iDeel = request.GET.get('deel', 1)
        iSectie = request.GET.get('sectie', None)
        iAflnum = request.GET.get('aflnum', 1)
        sFile = request.GET.get('filename', '')

        bUseDbase = request.GET.get('usedbase', False)
        if bUseDbase:
            if bUseDbase == "true":
                bUseDbase = True
            else:
                bUseDbase = False

        # Get the id of the Info object
        if iSectie==None or iSectie == "":
            info = Info.objects.filter(deel=iDeel, aflnum=iAflnum).first()
        else:
            info = Info.objects.filter(deel=iDeel, sectie=iSectie, aflnum=iAflnum).first()

        if info == None:
            data['status'] = 'error: no Info object found'
            return JsonResponse(data)

        # Remove any previous status objects for this info
        Status.objects.filter(info=info).delete()

        # Create a new import-status object
        oStatus = Status(info=info)

        # Note that we are starting
        oStatus.set_status("starting")
        iStatus = oStatus.id
        # oCsvImport['status'] = "starting"

        # Call the process - depending on the file ending
        if ".xml" in sFile:
            oResult = xml_to_fixture(sFile, iDeel, iSectie, iAflnum, iStatus, bUseDbase = bUseDbase, bUseOld = True)
        else:
            oResult = csv_to_fixture(sFile, iDeel, iSectie, iAflnum, iStatus, bUseDbase = bUseDbase, bUseOld = True)
        if oResult == None or oResult['result'] == False:
            data['status'] = 'error'

        # WSince we are done: explicitly set the status so
        oStatus.set_status("done")
    except Exception as ex:
        oErr.DoError("import_csv_start error")
        data['status'] = "error"

    # Return this response
    return JsonResponse(data)

def import_csv_progress(request):
    oErr = ErrHandle()
    # Prepare a return object
    data = {'read':0, 'skipped':0, 'method': '(unknown)', 'msg': ''}
    try:
        # Debugging
        oErr.Status("import_csv_progress at {}".format(datetime.now()))

        if request.POST:
            qd = request.POST
        else:
            qd = request.GET
        iDeel = qd.get('deel', 1)
        iSectie = qd.get('sectie', None)
        iAflnum = qd.get('aflnum', 1)
        # Get the id of the Info object
        if iSectie==None or iSectie == "":
            info = Info.objects.filter(deel=iDeel, aflnum=iAflnum).first()
        else:
            info = Info.objects.filter(deel=iDeel, sectie=iSectie, aflnum=iAflnum).first()
        # Find out how far importing is going
        qs = Status.objects.filter(info=info)
        if qs != None and len(qs) > 0:
            oStatus = qs[0]
            # Fill in the return object
            data['read'] = oStatus.read
            data['skipped'] = oStatus.skipped
            data['method'] = oStatus.method
            data['status'] = oStatus.status
            # Checking...
            if data['status'] == "idle":
                data['msg'] = "Idle status in import_csv_progress"
        else:
            # Do we have an INFO object?
            if info == None:
                data['status'] = "Please supply [deel], [sectie] and [aflnum]"
            else:
                data['status'] = "No status object for info=" + str(info.id) + " has been created yet"
    except Exception as ex:
        oErr.DoError("import_csv_progress error")
        data['status'] = "error"

    # Return where we are
    return JsonResponse(data)

def import_update_start(request, pk):
    # Formulate a response
    data = {'status': 'done'}
    oErr = ErrHandle()

    try:
        sFile = request.GET.get('filename', '')

        bUseDbase = request.GET.get('usedbase', False)
        if bUseDbase:
            if bUseDbase == "true":
                bUseDbase = True
            else:
                bUseDbase = False

        # Get the id of the DataUpdate object
        dataupdate = DataUpdate.objects.filter(id=pk).first()

        if dataupdate == None:
            data['status'] = 'error: no DataUpdate object found'
            return JsonResponse(data)

        # Remove any previous status objects for this info
        StatusUpdate.objects.filter(dataupdate=dataupdate).delete()

        # Create a new import-status object
        oStatus = StatusUpdate.objects.create(dataupdate=dataupdate)

        # Note that we are starting
        oStatus.set_status("starting")

        # Call the process - depending on the file ending
        if ".xml" in sFile:
            oResult = xml_update(sFile, oStatus.id, dataupdate.id, bUseDbase = bUseDbase, bUseOld = True)
        else:
            # *NOT implemented* yet for CSV
            # oResult = csv_to_fixture(sFile, iDeel, iSectie, iAflnum, iStatus, bUseDbase = bUseDbase, bUseOld = True)
            pass

        if oResult == None or oResult['result'] == False:
            data['status'] = 'error'

        # WSince we are done: explicitly set the status so
        oStatus.set_status("done")
    except Exception as ex:
        oErr.DoError("import_update_start error")
        data['status'] = "error"

    # Return this response
    return JsonResponse(data)

def import_update_progress(request, pk):
    oErr = ErrHandle()
    # Prepare a return object
    data = {'read':0, 'skipped':0, 'method': '(unknown)', 'msg': ''}
    try:
        # Debugging
        oErr.Status("import_update_progress at {}".format(datetime.now()))

        qd = request.POST if request.POST else request.GET
        # Get the DataUpdate object
        dataupdate = DataUpdate.objects.filter(id=pk).first()
        # Find out how far importing is going
        qs = StatusUpdate.objects.filter(dataupdate=dataupdate)
        if qs != None and len(qs) > 0:
            oStatus = qs[0]
            # Fill in the return object
            data['read'] = oStatus.read
            data['skipped'] = oStatus.skipped
            data['method'] = oStatus.method
            data['status'] = oStatus.status
            # Checking...
            if data['status'] == "idle":
                data['msg'] = "Idle status in import_update_progress"
        else:
            # Do we have an INFO object?
            if dataupdate == None:
                data['status'] = "Cannot find DataUpdate with id={}".format(pk)
            else:
                data['status'] = "No status object for dataupdate=" + str(dataupdate.id) + " has been created yet"
    except Exception as ex:
        oErr.DoError("import_update_progress error")
        data['status'] = "error"

    # Return where we are
    return JsonResponse(data)

def import_kloeke_cumul():
    """Import kloeke information from the DSDD file
    
    Column significance:
         0 kloeke_code1        De huidige kloekecode
         1 kloeke_code_oud     (deze kloekecodes gebruiken we niet meer)
         2 plaats
         3 gemeentecode_2007   Code voor deze gemeente
         4 gemeente
         5 streek           
         6 provincie
         7 postcode            
         8 land                Land
         9 rd_x                Rijksdriehoeks X           
        10 rd_y                Rijksdriehoeks Y
        11 lat                 latitude
        12 lng                 longitude
        13 topocode            ?
        14 std_spelling        ?
        15 volgnr              nummer (intern)
        16 in_dsdd             boolean (false)
        17 cannot_be_in_wvd    boolean (false)
        18 doet_niet_mee       boolean (false)
        19 in_wvd              boolean (false)
    """

    file = "kloeke_cumul.tsv"
    oErr = ErrHandle()
    bSuccess = False
    try:
        # FIgure out where the file is actually to be found
        file = os.path.abspath(os.path.join(MEDIA_ROOT, file))
        # Read the file into memory
        with open(file, "r", encoding="utf-8") as fd:
            lKloekeInfo = fd.readlines() # json.load(fd)

        with transaction.atomic():
            for tabline in lKloekeInfo:
                oInfo = tabline.replace('\n', '').split('\t')
                # Each item contains 5 elements: id, kloeke, place, x, y
                kloeke = oInfo[0]
                province = oInfo[6]
                country = oInfo[8]
                place = oInfo[2]
                point = '{}, {}'.format(oInfo[11], oInfo[12])
                obj = Coordinate.objects.filter(kloeke__iexact=kloeke).first()
                if obj == None:
                    obj = Coordinate.objects.create(kloeke=kloeke, 
                        place=place, point=point, province=province, country=country)
                else:
                    bNeedSaving = False
                    # Adapt stuff
                    if obj.kloeke != kloeke: 
                        obj.kloeke = kloeke
                        bNeedSaving = True
                    if obj.province != province: 
                        obj.province = province
                        bNeedSaving = True
                    if obj.country != country: 
                        obj.country = country
                        bNeedSaving = True
                    if obj.place != place: 
                        obj.place = place
                        bNeedSaving = True
                    if obj.point != point: 
                        obj.point = point
                        bNeedSaving = True
                    if bNeedSaving:
                        obj.save()

                # Check if the link from [Dialect] has already been made
                dialect = Dialect.objects.filter(nieuw__iexact=kloeke).first()
                if dialect != None:
                    if dialect.coordinate == None:
                        dialect.coordinate = obj
                        dialect.save()

        bSuccess = True
    except:
        msg = oErr.get_error_message()
        oErr.DoError("import_kloeke_cumul")
        bSuccess = False
    return bSuccess


class DictionaryDetailView(DetailView):
    """Details of an entry from the dictionary"""

    model = Entry
    context_object_name = 'entry'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DictionaryDetailView, self).get_context_data(**kwargs)

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} detail".format(THIS_DICTIONARY)

        # Return the calculated context
        return context


class TrefwoordListView(ListView):
    """ListView of keywords (trefwoorden)"""

    model = Trefwoord
    template_name = 'dictionary/trefwoord_list.html'
    # paginate_by = paginateEntries
    paginate_by = paginateSize
    entrycount = 0
    bUseMijnen = False  # Limburg uses mijnen, Brabant not
    bWbdApproach = True # Filter using the WBD approach
    qEntry = None
    qs = None
    bDoTime = True      # Measure time
    strict = True       # Use strict filtering

    def get_qs(self):
        if self.qEntry == None:
            if self.qs != None:
                qs = self.qs
            else:
                # Get the PKs of Entry related to Trefwoord
                qs = self.get_queryset()
            # Get the Entry queryset related to this
            qs = Entry.objects.filter(trefwoord__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            return export_csv(self.get_qs(), 'trefwoorden')
        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""
            return export_xlsx(self.get_qs(), 'trefwoorden')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""
            return export_html(self.get_qs(), 'trefwoorden')
        else:
            oResponse = super(TrefwoordListView, self).render_to_response(context, **response_kwargs)
            return oResponse

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(TrefwoordListView, self).get_context_data(**kwargs)

        # Action depends on the approach
        if self.bWbdApproach:
            # Need to adapt the object_list to get the entries to be used
            context['object_list'] = self.get_entryset(context['page_obj'])

        # Get parameters for the search
        initial = self.request.GET
        if initial is None:
            initial = {'optdialect': 'stad'}

        # Fill the 'searchform' context variable with the values that are received from the GET request
        search_form = TrefwoordSearchForm(initial)
        context['searchform'] = search_form

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
            self.paginate_by = int(initial['paginate_by'])
        else:
            context['paginateSize'] = self.paginate_by  # paginateSize

        if self.bUseMijnen:
            # Try to retain the choice for Mijn
            if 'mijn' in initial:
                mijn_id = int(initial['mijn'])
                context['mijnkeuze'] = mijn_id
                mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                if mijn_inst == None:
                    context['mijnnaam'] = ''
                else:
                    context['mijnnaam'] = mijn_inst.naam
            else:
                context['mijnkeuze'] = 0
                context['mijnnaam'] = ''
        # Process and retain the choice for Aflevering
        if 'aflevering' in initial:
            context['aflkeuze'] = int(initial['aflevering'])
            afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
            if afl == None:
                context['afl'] = ''
            else:
                context['afl'] = afl.get_summary()
        else:
            context['aflkeuze'] = 0
            context['afl'] = ''

        # Get possible user choice of 'strict'
        if 'strict' in initial:
            self.strict = (initial['strict'] == "True")
        context['strict'] = str(self.strict)
        
        # Determine the count 
        context['entrycount'] = self.entrycount   #  self.get_queryset().count()
        # context['twcount'] = self.twcount

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the afleveringen and mijnen that are available
        context['afleveringen'] = [afl for afl in Aflevering.objects.all()]

        mijnen = []
        if self.bUseMijnen:
            mijnen = [mijn for mijn in Mijn.objects.all().order_by('naam')]
        context['mijnen'] = mijnen

        # Set the title of the application
        context['title'] = "{} trefwoorden".format(THIS_DICTIONARY)

        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
        context['year'] = datetime.now().year

        # If we are in 'strict' mode, we need to deliver the [qlist]
        if self.strict:
            # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
            lAflev = self.get_qafl(context)

            # Get a list with 'first' and 'last' values for each item in the current paginated queryset
            lEntry = self.get_qlist(context)
            # Add the sorted-dialect information to lEntry
            for idx, item in enumerate(lEntry):
                # Start or Finish dialect information
                if item['trefwoord_woord']['first']:
                    qsa = []
                # All: add this entry
                qsa.append(lAflev[idx])
                if item['trefwoord_woord']['last']:
                    # COpy the list of Entry elements sorted by Trefwoord/Aflevering here
                    lEntry[idx]['alist'] = qsa
                else:
                    lEntry[idx]['alist'] = None

            context['qlist'] = lEntry

        # Return the calculated context
        return context
      
    def get_qlist(self, context):
        """Get a list of Entry elements + first/last information"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        lVars = ["trefwoord_woord", "lemma_gloss", "toelichting", "dialectopgave", "dialect_stad"]
        lFuns = [["trefwoord", "woord"], ["lemma", "gloss"], Entry.get_toelichting, Entry.dialectopgave, ["dialect", "stad"]]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Trefwoord/Aflevering into a list"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        qsd = []
        # Walk through the query set
        for entry in qs:
            qsd.append(entry)
        # Now sort the resulting set
        qsd = sorted(qsd, key=lambda el: el.trefwoord.woord + " " + el.get_aflevering())
        # Prepare for processing
        lVarsD = ["trefw", "afl"]
        lFunsD = [["trefwoord", "woord"], Entry.get_aflevering]
        # Create a list of the items
        lAfl = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lAfl            
      
    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)

    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Retrieve the set of trefwoorden from the page_obj
        trefw_list = [item.id for item in page_obj]

        # Initialize the filtering
        lstQ.append(Q(trefwoord__id__in=trefw_list))

        # Get the parameters passed on with the GET request
        get = self.request.GET

        # Check for aflevering being publishable
        if self.strict:
            lstQ.append(Q(aflevering__toonbaar=True))
        else:
            lstQ.append(Q(entry__aflevering__toonbaar=True))

        # Check for dialectwoord
        if 'dialectwoord' in get and get['dialectwoord'] != '':
            val = adapt_search(get['dialectwoord'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(woord__iregex=val))
            else:
                lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for lemma
        if 'lemma' in get and get['lemma'] != '':
            val = adapt_search(get['lemma'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(lemma__gloss__iregex=val))
            else:
                lstQ.append(Q(entry__lemma__gloss__iregex=val))
            bHasFilter = True

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(dialect__stad__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            # Adapt Entry filter
            if self.strict:
                lstQ.append(Q(dialect__nieuw__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(aflevering__id=iVal))
                    else:
                        lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(mijnlijst__id=iVal))
                    else:
                        lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Order: "trefwoord_woord", "lemma_gloss", "dialectopgave", "dialect_stad"
        # Make sure we apply the filter
        qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
            Lower('trefwoord__woord'), 
            'lemma__gloss',  
            Lower('toelichting'),
            Lower('woord'), 
            Lower('dialect__stad'))
        self.qEntry = qse
        return qse
        
    def get_queryset(self):

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        
        # Debugging: mesaure time
        if self.bDoTime: iStart = get_now_time()

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = get['search']
            if '*' in val or '[' in val or '?' in val or '#' in val:
                val = adapt_search(val)
                lstQ.append(Q(woord__iregex=val) )
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(woord__iexact=val))
            #val = adapt_search(get['search'])
            ## Use the 'woord' attribute of Trefwoord
            #lstQ.append(Q(woord__iregex=val) )
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                lstQ.append(Q(sn__exact=val))

        # Check for 'toelichting'
        if 'toelichting' in get and get['toelichting'] != '':
            val = adapt_search(get['toelichting'])
            # Try to get to the 'toelichting'
            lstQ.append(Q(toelichting__iregex=val))
            bHasSearch = True

        # Check for dialectwoord
        if 'dialectwoord' in get and get['dialectwoord'] != '':
            val = adapt_search(get['dialectwoord'])
            # Adapt Entry filter
            lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for lemma
        if 'lemma' in get and get['lemma'] != '':
            val = adapt_search(get['lemma'])
            # Adapt Entry filter
            lstQ.append(Q(entry__lemma__gloss__iregex=val))
            bHasFilter = True

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            # Adapt Entry filter
            lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            # Adapt Entry filter
            lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Debugging: time
        if self.bDoTime: print("TrefwoordListView get_queryset part 1: {:.1f}".format(get_now_time() - iStart))

        # Debugging: mesaure time
        if self.bDoTime: iStart = get_now_time()

        # Figure out which trefwoorden to exclude
        trefwoord_exclude = Trefwoord.objects.filter(toonbaar=0)

        # Create a QSE
        qse = Trefwoord.objects.exclude(id__in=trefwoord_exclude).filter(*lstQ).select_related().order_by(Lower('woord')).distinct()

        # Debugging: time
        if self.bDoTime: 
                print("TrefwoordListView get_queryset part 2: {:.1f}".format(get_now_time() - iStart))
                iStart = get_now_time()

        # Note the number of ITEMS we have
        #   (The nature of these items depends on the approach taken)
        # self.entrycount = qse.count()
        # Using 'len' is faster since [qse] is being actually used again
        self.entrycount = len(qse)

        # Debugging: time
        if self.bDoTime: 
                print("TrefwoordListView get_queryset part 3: {:.1f}".format(get_now_time() - iStart))

        return qse


class LemmaListView(ListView):
    """ListView of lemma's"""

    model = Lemma
    # context_object_name = 'lemma'    
    template_name = 'dictionary/lemma_list.html'
    template_ajax = 'dictionary/lemma_list_oview.html'
    # paginate_by = paginateEntries
    paginate_by = paginateSize
    entrycount = 0
    bUseMijnen = False      # Limburg uses mijnen, Brabant not
    bWbdApproach = True     # Filter using the WBD approach
    bOrderWrdToel = False   # Use the word order 'dialectopgave-toelichting' if True
    bDoTime = True          # Measure time
    qEntry = None
    qs = None
    strict = True      # Use strict filtering ALWAYS

    def get_qs(self):
        """Get the Entry elements that are selected"""
        if self.qEntry == None:
            # Get the [Lemma] pk's that are relevant
            if self.qs != None:
                qs = self.qs
            else:
                # Get the Lemma PKs
                qs = self.get_queryset()
            # Get the Entry queryset related to this
            qs = Entry.objects.filter(lemma__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            
            return export_csv(self.get_qs(), 'begrippen')
        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""

            return export_xlsx(self.get_qs(), 'begrippen')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""

            return export_html(self.get_qs(), 'begrippen')

        else:
            iStart = get_now_time()
            # sResp = render_to_string(self.template_name, context)
            oRendered = super(LemmaListView, self).render_to_response(context, **response_kwargs)
            # Show what the render-time was
            print("LemmaListView render time: {:.1f}".format(get_now_time() - iStart))
            return oRendered

    def post(self, request, *args, **kwargs):
        # Prepare return object
        oData = {'status': 'error', 'msg': ''}
        oErr = ErrHandle()
        try:
            self.object_list = self.get_queryset()
            allow_empty = self.get_allow_empty()

            if not allow_empty:
                # When pagination is enabled and object_list is a queryset,
                # it's better to do a cheap query than to load the unpaginated
                # queryset in memory.
                if self.get_paginate_by(self.object_list) is not None and hasattr(self.object_list, 'exists'):
                    is_empty = not self.object_list.exists()
                else:
                    is_empty = len(self.object_list) == 0
                if is_empty:
                    raise Http404(_("Empty list and '%(class_name)s.allow_empty' is False.") % {
                        'class_name': self.__class__.__name__,
                    })
            # Start collecting context time
            if self.bDoTime: iStart = get_now_time()

            context = self.get_context_data()
            if self.bDoTime:
                print("LemmaListView context [a]: {:.1f}".format( get_now_time() - iStart))
                iStart = get_now_time()

            sText = render_to_string(self.template_ajax, context, request)
            if self.bDoTime:
                print("LemmaListView context [b]: {:.1f}".format( get_now_time() - iStart))

            oData['html'] = sText
            oData['status'] = "ok"
        except:
            oData['msg'] = oErr.get_error_message()
        return JsonResponse(oData)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(LemmaListView, self).get_context_data(**kwargs)

        oErr = ErrHandle()
        try:
            # Get parameters for the search
            # initial = self.request.GET
            initial = self.get
            if initial is None:
                initial = {'optdialect': 'stad'}

            # Possibly set the page object
            if 'page' in initial:
                oPage = context['page_obj']
                oPage.number = int(initial['page'])
                context['page_obj'] = oPage

            # Action depends on the approach
            if self.bWbdApproach:
                # Start collecting context time
                if self.bDoTime: iStart = get_now_time()
                # Need to adapt the object_list to get the entries to be used
                context['object_list'] = list(self.get_entryset(context['page_obj']))
                if self.bDoTime:
                    print("LemmaListView get_context_data: {:.1f}".format( get_now_time() - iStart))

            # Start collecting context time
            if self.bDoTime: iStart = get_now_time()

            search_form = LemmaSearchForm(initial)

            context['searchform'] = search_form

            # Get possible user choice of 'strict'
            if 'strict' in initial:
                self.strict = (initial['strict'] == "True")
            context['strict'] = str(self.strict)
        
            # Determine the count 
            context['entrycount'] = self.entrycount #  self.get_queryset().count()

            # Make sure the paginate-values are available
            context['paginateValues'] = paginateValues

            if 'paginate_by' in initial:
                context['paginateSize'] = int(initial['paginate_by'])
                self.paginate_by = int(initial['paginate_by'])
            else:
                context['paginateSize'] = self.paginate_by

            if self.bUseMijnen:
                # Try to retain the choice for Mijn
                if 'mijn' in initial:
                    mijn_id = int(initial['mijn'])
                    context['mijnkeuze'] = mijn_id
                    mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                    if mijn_inst == None:
                        context['mijnnaam'] = ''
                    else:
                        context['mijnnaam'] = mijn_inst.naam
                else:
                    context['mijnkeuze'] = 0
                    context['mijnnaam'] = ''
            # Process and retain the choice for Aflevering
            if 'aflevering' in initial and initial['aflevering'] != "":
                context['aflkeuze'] = int(initial['aflevering'])
                afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
                if afl == None:
                    context['afl'] = ''
                else:
                    context['afl'] = afl.get_summary()
            else:
                context['aflkeuze'] = 0
                context['afl'] = ''

            # Set the prefix
            context['app_prefix'] = APP_PREFIX

            # Set the title of the application
            context['title'] = "{} begrippen".format(THIS_DICTIONARY)

            # Set the afleveringen that are available
            context['afleveringen'] = [afl for afl in Aflevering.objects.all()]

            mijnen = []
            if self.bUseMijnen:
                mijnen = [mijn for mijn in Mijn.objects.all().order_by('naam')]
            context['mijnen'] = mijnen

            # Pass on the word-order boolean
            context['order_word_toel'] = self.bOrderWrdToel

            if self.bDoTime:
                print("LemmaListView context part 1: {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # If we are in 'strict' mode, we need to deliver the [qlist] and its subsidiaries
            if self.strict:

                # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
                lAflev = self.get_qafl(context)
                if self.bDoTime:
                    print("LemmaListView context get_qafl(): {:.1f}".format( get_now_time() - iStart))
                    # Reset the time
                    iStart = get_now_time()

                lastDescr = None
                # Get a list with 'first' and 'last' values for each item in the current paginated queryset
                lEntry = self.get_qlist(context)
                if self.bDoTime:
                    print("LemmaListView context get_qlist(): {:.1f}".format( get_now_time() - iStart))
                    # Reset the time
                    iStart = get_now_time()

                # Add the sorted-dialect information to lEntry
                for idx, item in enumerate(lEntry):
                    # Start or Finish dialect information
                    if item['lemma_gloss']['first']:
                        qsa = []
                        # start a list of all lemma-descriptions taken from the entries belonging to one particular lemma
                        lemma_descr_list = []
                    # All: add this entry
                    qsa.append(lAflev[idx])
                    # If not already present: add description
                    iDescrId = item['entry'].descr.id
                    if iDescrId not in lemma_descr_list:
                        lemma_descr_list.append(iDescrId)
                    if item['lemma_gloss']['last']:
                        # COpy the list of Entry elements sorted by Lemma/Aflevering here
                        lEntry[idx]['alist'] = qsa
                        # OLD: lEntry[idx]['dlist'] = self.get_qdescr(item['entry'])
                        lEntry[idx]['dlist'] = self.get_qdescrlist(lemma_descr_list)
                    else:
                        lEntry[idx]['alist'] = None
                        lEntry[idx]['dlist'] = None

                context['qlist'] = lEntry
            # Finish measuring context time
            if self.bDoTime:
                print("LemmaListView context: {:.1f}".format( get_now_time() - iStart))

            # Set the method
            context['method'] = "get"       # Alternative: "ajax"

            context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
            context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
            context['year'] = datetime.now().year
        except:
            msg = oErr.get_error_message()
            oErr.DoError("LemmaListView/get_context_data")

        # Return the calculated context
        return context

    def get_qlist(self, context):
        """Get a list of Entry elements + first/last information"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        if self.bOrderWrdToel:
            lVars = ["lemma_gloss", "trefwoord_woord", "dialectopgave", "toelichting", "dialect_stad"]
            lFuns = [["lemma", "gloss"], ["trefwoord", "woord"], Entry.dialectopgave, Entry.get_toelichting, ["dialect", "stad"]]
        else:
            lVars = ["lemma_gloss", "trefwoord_woord", "toelichting", "dialectopgave", "dialect_stad"]
            lFuns = [["lemma", "gloss"], ["trefwoord", "woord"], Entry.get_toelichting, Entry.dialectopgave, ["dialect", "stad"]]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Lemma/Aflevering into a list"""

        bMethodQset = False

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']

        if bMethodQset:
            # Create a list of entry ids
            id_list = [item.id for item in qs]

            # Create the correctly sorted queryset
            qsd = Entry.objects.filter(Q(id__in=id_list)).select_related().order_by(
                'lemma__gloss', 
                'aflevering__deel__nummer', 
                'aflevering__sectie', 
                'aflevering__aflnum')
        else:
            # qsd = copy.copy(qs)
            # Force evaluation
            qsd = list(qs)
            # Now sort the resulting set
            qsd = sorted(qsd, key=lambda el: (el.lemma.gloss, el.get_aflevering()) )

        # Prepare for processing
        lVarsD = ["lem", "afl"]
        lFunsD = [["lemma", "gloss"], Entry.get_aflevering]
        # Create a list of the items
        lAfl = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lAfl            

    def get_qdescr(self, entry):
        """OLD: Sort the paginated QS by Lemma/lmdescr.toelichting into a list"""

        # Get the sorted set of [lmdescr] objects for this lemma
        #qsd = []
        ## Walk through the query set
        #for d in entry.lemma.lmdescr.order_by(Lower('toelichting'), Lower('bronnenlijst')): qsd.append(d)
        qsd = [d for d in entry.lemma.lmdescr.order_by(Lower('toelichting'), 'bronnenlijst')]
        # Prepare for processing
        lVarsD = ["descr", "bronnen"]
        lFunsD = [["toelichting"], ["bronnenlijst"]]
        # Create a list of the items
        lDescr = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDescr            
      
    def get_qdescrlist(self, descr_id_list):
        """Sort the paginated QS by Lemma/lmdescr.toelichting into a list"""

        # Make a query that gets the indicated id's
        qsd = Description.objects.filter(Q(id__in=descr_id_list)).order_by(Lower('toelichting'), Lower('bronnenlijst'))
        # Prepare for processing
        lVarsD = ["descr", "bronnen"]
        lFunsD = [["toelichting"], ["bronnenlijst"]]
        # Create a list of the items
        lDescr = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDescr            
      
    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Initialize timer
        if self.bDoTime: iStart = get_now_time()

        # Retrieve the set of trefwoorden from the page_obj
        lemma_list = [item.id for item in page_obj.object_list]
        if self.bDoTime: print("LemmaListView get_entryset part 1: {:.1f}".format(get_now_time() - iStart))

        if self.bDoTime: iStart = get_now_time()
        # Initialize the filtering
        lstQ.append(Q(lemma__id__in=lemma_list))
        # lstQ.append(Q(lemma__id__in=page_obj))

        # Get the parameters passed on with the GET request
        get = self.get

        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = adapt_search(get['dialectCity'])
            if self.strict:
                lstQ.append(Q(dialect__stad__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = adapt_search(get['dialectCode'])
            if self.strict:
                lstQ.append(Q(dialect__nieuw__iregex=val))
            else:
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            bHasFilter = True

        # Check for dialect word, which is a direct member of Entry
        if 'woord' in get and get['woord'] != '':
            val = adapt_search(get['woord'])
            if self.strict:
                lstQ.append(Q(woord__iregex=val))
            else:
                lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(aflevering__id=iVal))
                    else:
                        lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    if self.strict:
                        lstQ.append(Q(mijnlijst__id=iVal))
                    else:
                        lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Make sure we filter on aflevering.toonbaar
        if self.strict:
            lstQ.append(Q(aflevering__toonbaar=True))
        else:
            lstQ.append(Q(entry__aflevering__toonbaar=True))

        if self.bDoTime: print("LemmaListView get_entryset part 2: {:.1f}".format(get_now_time() - iStart))

        # Make the QSE available
        # Order: "lemma_gloss", "trefwoord_woord", "dialectopgave", "dialect_stad"
        if self.bDoTime: iStart = get_now_time()
        if self.bOrderWrdToel:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('lemma__gloss'),  
                Lower('trefwoord__woord'), 
                Lower('woord'), 
                Lower('toelichting'), 
                Lower('dialect__stad'))
        else:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('lemma__gloss'),  
                Lower('trefwoord__woord'), 
                Lower('toelichting'), 
                Lower('woord'), 
                Lower('dialect__stad'))
        if self.bDoTime: print("LemmaListView get_entryset part 3: {:.1f}".format(get_now_time() - iStart))
        # x = str(Entry.objects.filter(*lstQ).distinct().select_related().query)
        self.qEntry = qse
        return qse

    def get_queryset(self):
        # Measure how long it takes
        if self.bDoTime:
            iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        self.get = get

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = get['search']
            if '*' in val or '[' in val or '?' in val or '#' in val:
                val = adapt_search(val)
                lstQ.append(Q(gloss__iregex=val) )
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(gloss__iexact=val))
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                lstQ.append(Q(sn__exact=val))
 
        # Check for dialect city
        if 'dialectCity' in get and get['dialectCity'] != '':
            val = get['dialectCity']
            if '*' in val or '[' in val or '?' in val or '#' in val:
                # val = adapt_search(get['dialectCity'])
                val = adapt_search(val)
                lstQ.append(Q(entry__dialect__stad__iregex=val))
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(entry__dialect__stad__iexact=val))
            bHasFilter = True

        # Check for dialect code (Kloeke)
        if 'dialectCode' in get and get['dialectCode'] != '':
            val = get['dialectCode']
            if '*' in val or '[' in val or '?' in val or '#' in val:
                val = adapt_search(val)
                lstQ.append(Q(entry__dialect__nieuw__iregex=val))
            else:
                # Strive for equality, but disregard case
                lstQ.append(Q(entry__dialect__nieuw__iexact=val))
            bHasFilter = True

        # Check for dialect word, which is a direct member of Entry
        if 'woord' in get and get['woord'] != '':
            val = adapt_search(get['woord'])
            lstQ.append(Q(entry__woord__iregex=val))
            bHasFilter = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal))
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and  'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal))
                    bHasFilter = True

        # Method #8 -- use the lemma.toonbaar property
        lemma_hide = Lemma.objects.filter(toonbaar=0)

        qse = Lemma.objects.exclude(id__in=lemma_hide).filter(*lstQ).select_related().order_by('gloss').distinct()

        # Time measurement
        if self.bDoTime:
            print("LemmaListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            print("LemmaListView query: {}".format(qse.query))

        # Note the number of ITEMS we have
        #   (The nature of these items depends on the approach taken)
        # self.entrycount = qse.count()
        # Note: while taking more time here, it saves time later
        self.entrycount = len(qse)

        # Time measurement
        if self.bDoTime:
            print("LemmaListView get_queryset point 'b': {:.1f}".format( get_now_time() - iStart))

        # Return the resulting filtered and sorted queryset
        return qse


class LemmaMapView(MapView):
    model = Lemma
    modEntry = Entry
    frmSearch = LemmaSearchForm
    order_by = ["trefwoord"]
    labelfield = "gloss"

    def initialize(self):
        super(LemmaMapView, self).initialize()
        # Entries with a 'form' value
        self.add_entry('woord', 'str', 'woord', 'woord')
        self.add_entry('stad', 'str', 'dialect__stad', 'dialectCity')
        self.add_entry('kloeke', 'str', 'dialect__nieuw', 'dialectCode')
        self.add_entry('aflevering', 'int', 'aflevering__id', 'aflevering')
        self.add_entry('mijn', 'int', 'mijnlijst__id', 'mijn')

        # Entries without a 'form' value
        self.add_entry('trefwoord', 'str', 'trefwoord__woord')
        self.add_entry('point', 'str', 'dialect__coordinate__point')
        self.add_entry('place', 'str', 'dialect__coordinate__place')

    def get_popup(self, entry):
        """Create a popup from the 'key' values defined in [initialize()]"""

        pop_up = '<p class="h6">{}</p>'.format(entry['woord'])
        pop_up += '<hr style="border: 1px solid green" />'
        pop_up += '<p style="font-size: smaller;"><span style="color: purple;">{}</span> {}</p>'.format(
            entry['kloeke'], entry['stad'])
        return pop_up
    

class LocationListView(ListView):
    """Listview of locations"""

    model = Dialect     # The LocationListView uses [Dialect]
    template_name = 'dictionary/location_list.html'
    # paginate_by = paginateEntries # paginateSize
    paginate_by = paginateLocations    # Default pagination number SPECIFICALLY for dialects (1)
    entrycount = 0      # Number of items in queryset (whether Entry or Dialect!!)
    bUseMijnen = False  # Limburg uses mijnen, Brabant not, Gelderland neither
    bWbdApproach = True # Filter using the WBD approach
    bNewOrder = True    # Use the new order (see issue #22 of the RU-wld)
    qEntry = None       # Current queryset as restricted to ENTRY
    qAll = None         # Ordered queryset of ALL
    qs = None           # Current queryset (for speeding up)
    strict = True       # Use strict filtering ALWAYS
    bDoTime = True      # Use timing to determine what goes fastest

    def get_qs(self):
        """Get the Entry elements that are selected"""

        if self.qEntry == None:
            # First get the currently selected elements
            if self.qs != None:
                qs = self.qs
            else:
                # Calculate the  PKs
                qs = self.get_queryset()
            # Get the Entry elements that refer to the set of dialects
            if not self.strict:
                # Convert the [Dialect] elements in qs to [Entry] elements
                qs = Entry.objects.filter(dialect__pk__in=qs).select_related()
        else:
            qs = self.qEntry
        return qs

    def render_to_response(self, context, **response_kwargs):
        """Check if a CSV response is needed or not"""
        if 'Csv' in self.request.GET.get('submit_type', ''):
            """ Provide CSV response"""
            return export_csv(self.get_qs(), 'plaatsen')

        elif 'Excel' in self.request.GET.get('submit_type', ''):
            """ Provide Excel response"""
            return export_xlsx(self.get_qs(), 'plaatsen')
        elif 'Html' in self.request.GET.get('submit_type', ''):
            """ Provide Html response"""
            return export_html(self.get_qs(), 'plaatsen')
        else:
            return super(LocationListView, self).render_to_response(context, **response_kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(LocationListView, self).get_context_data(**kwargs)

        # Start collecting context time
        if self.bDoTime: iStart = get_now_time()

        # Action depends on the approach
        if self.bWbdApproach:
            # Need to adapt the object_list to get the entries to be used
            # context['object_list'] = self.get_entryset(context['page_obj'])
            context['object_list'] = list(self.get_entryset(context['page_obj']))
            if self.bDoTime:
                print("LocationListView context - get_entryset(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

        # Get parameters for the search
        initial = self.request.GET
        search_form = DialectSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.entrycount   #  self.get_queryset().count()

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if self.bDoTime:
            print("LocationListView context part 0: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # Set the afleveringen that are available
        context['afleveringen'] = [afl for afl in Aflevering.objects.all()]

        mijnen = []
        if self.bUseMijnen:
            mijnen = [mijn for mijn in Mijn.objects.all().order_by('naam')]
        context['mijnen'] = mijnen

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
            self.paginate_by = int(initial['paginate_by'])
        else:
            context['paginateSize'] = self.paginate_by

        if self.bDoTime:
            print("LocationListView context part 1: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()

        # Try to retain the choice for Aflevering and Mijn
        if self.bUseMijnen:
            # Try to retain the choice for Mijn
            if 'mijn' in initial:
                mijn_id = int(initial['mijn'])
                context['mijnkeuze'] = mijn_id
                mijn_inst = Mijn.objects.filter(id=mijn_id).first()
                if mijn_inst == None:
                    context['mijnnaam'] = ''
                else:
                    context['mijnnaam'] = mijn_inst.naam
            else:
                context['mijnkeuze'] = 0
                context['mijnnaam'] = ''

        if 'aflevering' in initial:
            context['aflkeuze'] = int(initial['aflevering'])
            afl = Aflevering.objects.filter(id=context['aflkeuze']).first()
            if afl == None:
                context['afl'] = ''
            else:
                context['afl'] = afl.get_summary()
        else:
            context['aflkeuze'] = 0
            context['afl'] = ''

        if self.bDoTime:
            print("LocationListView context part 2: {:.1f}".format( get_now_time() - iStart))
            # Reset the time
            iStart = get_now_time()
            
        # Set the title of the application
        context['title'] = "{} plaatsen".format(THIS_DICTIONARY)

        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)

        # Get possible user choice of 'strict'
        if 'strict' in initial:
            self.strict = (initial['strict'] == "True")
        context['strict'] = str(self.strict)

        if self.strict:
            # Transform the paginated queryset into a dict sorted by Dialect/Aflevering
            lDialect = self.get_qafl(context)
            if self.bDoTime:
                print("LocationListView context get_qafl(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # Get a list with 'first' and 'last' values for each item in the current paginated queryset
            lEntry = self.get_qlist(context)

            if self.bDoTime:
                print("LocationListView context get_qlist(): {:.1f}".format( get_now_time() - iStart))
                # Reset the time
                iStart = get_now_time()

            # Add the sorted-dialect information to lEntry
            for idx, item in enumerate(lEntry):
                # Start or Finish dialect information
                if item['dialect_stad']['first']:
                    iDialectStart = idx
                    qsa = []
                # All: add this entry
                qsa.append(lDialect[idx])
                if item['dialect_stad']['last']:
                    # COpy the list of Entry elements sorted by Stad/Aflevering here
                    lEntry[idx]['alist'] = qsa
                else:
                    lEntry[idx]['alist'] = None

            context['qlist'] = lEntry

        # Finish measuring context time
        if self.bDoTime:
            print("LocationListView context: {:.1f}".format( get_now_time() - iStart))

        # Return the calculated context
        return context
      
    def get_qlist(self, context):
        """Calculate HTML output for the query-set in the context"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        # Start the output
        html = []
        # Initialize the variables whose changes are important
        lVars = ["dialect_stad", "lemma_gloss", "trefwoord_woord","toelichting", "dialectopgave"]
        lFuns = [["dialect", "stad"], ["lemma", "gloss"], ["trefwoord", "woord"], Entry.get_toelichting, Entry.dialectopgave]
        # Get a list of items containing 'first' and 'last' information
        lItem = get_item_list(lVars, lFuns, qs)
        # REturn this list
        return lItem

    def get_qafl(self, context):
        """Sort the paginated QS by Dialect/Aflevering and turn into a dict"""

        # REtrieve the correct queryset, as determined by paginate_by
        qs = context['object_list']
        qsd = list(qs)
        # Now sort the resulting set
        qsd = sorted(qsd, key=lambda el: el.dialect.stad + " " + el.get_aflevering())

        # Prepare for dialect processing
        lVarsD = ["stad", "afl"]
        lFunsD = [["dialect", "stad"], Entry.get_aflevering]
        # Create a list of Dialect items
        lDialect = get_item_list(lVarsD, lFunsD, qsd)
        # Return the result
        return lDialect            

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)

    def get_entryset(self, page_obj):
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Time measurement
        if self.bDoTime: iStart = get_now_time()

        # Initialize the filtering on ENTRY
        dialect_list = [item.id for item in page_obj]
        lstQ.append(Q(dialect__id__in=dialect_list))

        # Make sure we filter on aflevering.toonbaar
        lstQ.append(Q(aflevering__toonbaar=True))

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:a'): {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(aflevering__id=iVal) )
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(mijnlijst__id=iVal) )
                    bHasFilter = True

        bUseLower = True
        if bUseLower:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                Lower('dialect__stad'),
                'lemma__gloss',  
                Lower('trefwoord__woord'), 
                Lower('toelichting'), 
                Lower('woord'))
        else:
            qse = Entry.objects.filter(*lstQ).distinct().select_related().order_by(
                'dialect__stad',
                'lemma__gloss',  
                'trefwoord__woord', 
                'toelichting', 
                'woord')

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:b'): {:.1f}".format( get_now_time() - iStart))

        self.qEntry = qse

        qse = list(qse)
        # Time measurement
        if self.bDoTime:
            print("LocationListView get_entryset (point 'e:c'): {:.1f}".format( get_now_time() - iStart))

        return qse

    def get_queryset(self):
        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        # Not sure why, but get a copy
        get = get.copy()

        # Measure how long it takes
        if self.bDoTime:
            iStart = get_now_time()

        # Set the [sortOrder] parameter to 'stad' (the name of the city)
        get['sortOrder'] = 'stad'

        # Get possible user choice of 'strict'
        if 'strict' in get:
            self.strict = (get['strict'] == "True")

        # Queryset: build a list of requirements
        lstQ = []
        bHasSearch = False
        bHasFilter = False

        # Fine-tuning: search string is the STAD
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            lstQ.append(Q(stad__iregex=val) )
            bHasSearch = True

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)
                lstQ.append(Q(sn__exact=val))

        # Check for dialect code (Kloeke)
        if 'nieuw' in get and get['nieuw'] != '':
            val = adapt_search(get['nieuw'])
            lstQ.append(Q(nieuw__iregex=val) )
            bHasSearch = True

        # Check for aflevering
        if 'aflevering' in get and get['aflevering'] != '':
            # What we get should be a number
            val = get['aflevering']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__aflevering__id=iVal) )
                    bHasFilter = True

        # Check for mijn
        if self.bUseMijnen and 'mijn' in get and get['mijn'] != '':
            # What we get should be a number
            val = get['mijn']
            if val.isdigit():
                iVal = int(val)
                if iVal>0:
                    lstQ.append(Q(entry__mijnlijst__id=iVal) )
                    bHasFilter = True

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Get a list of Dialects that should be excluded
        dialect_hide = Dialect.objects.filter(toonbaar=0)

        # Use the E-WBD approach: be efficient here
        qs = Dialect.objects.exclude(id__in=dialect_hide).filter(*lstQ).distinct().select_related().order_by(Lower('stad'))

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'c': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # self.entrycount = qs.count()
        # Using 'len' is faster since [qse] is being actually used again
        self.entrycount = len(qs)

        # Time measurement
        if self.bDoTime:
            print("LocationListView get_queryset point 'd': {:.1f}".format( get_now_time() - iStart))
            iStart = get_now_time()

        # Return the resulting filtered and sorted queryset
        return qs


class DialectListView(ListView):
    """Listview of dialects"""

    model = Dialect
    paginate_by = 10
    template_name = 'dictionary/dialect_list.html'
    entrycount = 0
    bDoTime = True
    bImportKloekeInfo = False

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DialectListView, self).get_context_data(**kwargs)

        # One-time calls
        if self.bImportKloekeInfo: 
            import_kloeke_cumul()

        # Get parameters for the search
        initial = self.request.GET
        search_form = DialectSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.entrycount # self.get_queryset().count()

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Set the title of the application
        context['title'] = "{} dialecten".format(THIS_DICTIONARY)

        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)

        # Return the calculated context
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)

    def initialize(self):
        # Check if "Rhede" has been processed
        if Information.get_kvalue("rhede") != "done":
            adapt_dialect = []
            # Look for dialects that have the same name+streek, but multiple identifiers
            for dialect in Dialect.objects.all():
                stad = dialect.stad
                streek = dialect.streek
                count = Dialect.objects.filter(stad=stad, streek=streek).count()
                if count > 1:
                    bFound = False
                    for adaptation in adapt_dialect:
                        if stad == adaptation['stad'] and streek == adaptation['streek']:
                            bFound = True
                            break
                    if not bFound:
                        obj = Dialect.objects.filter(stad=stad, streek=streek).first()
                        all = [x.id for x in Dialect.objects.filter(stad=stad, streek=streek).all() if x.id != obj.id]
                        oAdapt = dict(stad=stad, streek=streek, obj=obj, all=all)
                        adapt_dialect.append(oAdapt)
            # Go through all the adaptations that need doing
            for adaptation in adapt_dialect:
                stad = adaptation['stad']
                streek = adaptation['streek']
                obj = adaptation['obj']
                all = adaptation['all']
                # Visit all entries with a dialect like this
                entries = Entry.objects.filter(dialect__id__in=all).values('id')
                with transaction.atomic():
                    for entry in Entry.objects.filter(id__in=entries):
                        entry.dialect = obj
                        entry.save()
                        print("done entry #{}".format(entry.id), file=sys.stderr)
            # Remove the superfluous dialects
            Dialect.objects.filter(id__in=all).delete()


            # Set the information value to 'done'
            Information.set_kvalue("rhede", "done")

        # Check if "DIalect count" has been processed
        if Information.get_kvalue("dialectcount") != "done":
            adapt_dialect = []
            # Visit all dialects and get the  number of entries per dialect
            with transaction.atomic():
                for dialect in Dialect.objects.all():
                    count = Entry.objects.filter(dialect=dialect).count()
                    dialect.count = count
                    dialect.save()


            # Set the information value to 'done'
            Information.set_kvalue("dialectcount", "done")

        return None
        
    def get_queryset(self):
        # Initializations
        self.initialize()

        # Measure how long it takes
        if self.bDoTime: iStart = get_now_time()

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        get = get.copy()
        self.get = get

        # Fix the sort-order
        get['sortOrder'] = 'stad'

        lstQ = []

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            query = Q(stad__iregex=val) 

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)

            # Apply the filter
            lstQ.append(query)

        # Check for dialect code (Kloeke)
        if 'nieuw' in get and get['nieuw'] != '':
            val = adapt_search(get['nieuw'])
            query = Q(nieuw__iregex=val)
            
            # Apply the filter
            lstQ.append(query)

        # Calculate the final qs
        qs = Dialect.objects.exclude(toonbaar=0).filter(*lstQ).order_by('stad').distinct()

        # Time measurement
        if self.bDoTime:
            print("DialectListView get_queryset point 'a': {:.1f}".format( get_now_time() - iStart))
            print("DialectListView query: {}".format(qs.query))
            iStart = get_now_time()

        # Determine the length
        self.entrycount = len(qs)

        # Time measurement
        if self.bDoTime:
            print("DialectListView get_queryset point 'b': {:.1f}".format( get_now_time() - iStart))

        # Return the resulting filtered and sorted queryset
        return qs


class DialectMapView(MapView):
    model = Dialect
    modEntry = Dialect
    frmSearch = DialectSearchForm
    order_by = ["streek", "stad"]
    use_object = False
    label = "Dialectplaatsen"

    def initialize(self):
        super(DialectMapView, self).initialize()
        # Entries with a 'form' value
        self.add_entry('stad', 'str', 'stad', 'search')
        self.add_entry('kloeke', 'str', 'nieuw', 'nieuw')

        # Entries without a 'form' value
        self.add_entry('trefwoord', 'str', 'streek')
        self.add_entry('point', 'str', 'coordinate__point')
        self.add_entry('place', 'str', 'coordinate__place')
        self.add_entry('count', 'int', 'count')

    def get_popup(self, dialect):
        """Create a popup from the 'key' values defined in [initialize()]"""

        pop_up = '<p class="h6">{}</p>'.format(dialect['stad'])
        pop_up += '<hr style="border: 1px solid green" />'
        pop_up += '<p style="font-size: smaller;"><span style="color: purple;">{}</span> {}</p>'.format(
            dialect['kloeke'], dialect['count'])
        return pop_up
    

class DialectCheckView(ListView):
    """Check how the dialects have fared"""

    model = Dialect
    template_name = 'dictionary/dialect_check.html'

    def get_context_data(self, **kwargs):
        # Retrieve the default context
        context = super(DialectCheckView, self).get_context_data(**kwargs)

        def get_range(qs, d, s):
            """Get the afleveringen belonging to d/s as a range-string"""

            lstQ = []
            lstQ.append(Q(deel__nummer=d))
            if s != None:
                lstQ.append(Q(sectie=s))
                sBack = "{}/{}:".format(d,s)
            else:
                sBack = "{}:".format(d)
            # qsr = qs.objects.filter(*lstQ).order_by('aflnum')
            # qsr = Aflevering.objects.filter(id__in=qs).filter(*lstQ).order_by('aflnum')
            qsr = qs.filter(*lstQ).order_by('aflnum')
            if qsr.count() == 0:
                return ""
            # Initializations
            prev_num = -1
            bRange = False  # In a range or not
            for item in qsr:
                num = item.aflnum
                if prev_num == -1:
                    sBack += "{}".format(num)
                elif prev_num + 1 == num:
                    # We are in a range
                    if not bRange:
                        # Indicate we are inside a range
                        sBack += "-"
                        bRange = True
                elif prev_num + 1 != num:
                    if bRange:
                        # Finish previous range and start a new one
                        sBack += "{},{}".format(prev_num, num)
                        bRange = False
                    else:
                        # Start a range
                        sBack += ",{}".format(num)
                # Adapt the previous number
                prev_num = num
            # Check if we are still in the range
            if bRange:
                # Finish this range
                sBack += "{}".format(num)
            return sBack

        # The set of ranges that cen be looked at
        ranges = [{'d': 1, 's': None}, {'d': 2, 's': None}, {'d': 3, 's': 1}, {'d': 3, 's': 2}, {'d': 3, 's': 3}, {'d': 3, 's': 4}]

        # Add my own stuff to the context
        # (1) Get a list of unique dialect names
        qs = Dialect.objects.order_by(Lower('stad')).distinct().select_related()
        d_list = []
        for item in qs:
            d_list.append({'id': item.id, 'stad': item.stad, 'nieuw': item.nieuw})
        context['d_list']  = d_list

        # (2) Get a list of unique kloekecodes
        qs = Dialect.objects.order_by(Lower('nieuw')).distinct().select_related()
        k_list = []
        for item in qs:
            k_list.append({'id': item.id, 'stad': item.stad, 'nieuw': item.nieuw})
        context['k_list']  = k_list


        # (3) Get a list of all names that have more than one entry (modula case)
        d_double = []
        last_stad = ""
        for item in d_list:
            name = item['stad']
            if name != last_stad:
                last_stad = name
                qs = Dialect.objects.filter(Q(stad__iexact=name))
                if qs.count()>1:
                    # Get all the codes and all the afl for this name
                    lCode = []
                    lAfl = []
                    for d in qs:
                        # Get all the entries that point to this dialect
                        qse = Entry.objects.filter(dialect=d).distinct()
                        qse_count = qse.count()
                        # Check the code (nieuw) for this dialect
                        if not d.nieuw in lCode:
                            lCode.append({'nieuw': d.nieuw,'num': qse_count})
                        # Check all the afl for this dialect
                        qsa = Aflevering.objects.filter(entry__in=qse).distinct()
                        for a in qsa:
                            if not a.id in lAfl: lAfl.append(a.id)
                    # Sort the list of afl
                    qsa = Aflevering.objects.filter(id__in=lAfl).select_related().order_by('deel', 'sectie', 'aflnum')
                    lAfl = []
                    # Get a number of standard ranges
                    for oRange in ranges:
                        sBack = get_range(qsa, oRange['d'], oRange['s'])
                        if sBack != "": lAfl.append(sBack)
                    # Get all the afleveringen for this name
                    oDouble = {'name': name, 'count': qs.count(), 'codes': lCode, 'afl_list': lAfl}
                    d_double.append(oDouble)
        context['d_double'] = d_double

        # (4) Get a list of all kloekecodes that have more than one entry 
        k_double = []
        last_code = ""
        for item in k_list:
            code = item['nieuw']
            if code != last_code:
                last_code = code
                # Get all the dialects with this particular kloekecode
                qs = Dialect.objects.filter(Q(nieuw__iexact=code)).distinct()
                if qs.count() > 1:
                    # There is more than one city linked to this kloekecode
                    lStad = []
                    lAfl = []
                    for d in qs:
                        # Get all the entries that point to this dialect
                        qse = Entry.objects.filter(dialect=d).distinct()
                        qse_count = qse.count()
                        # Check the city name (stad) for this dialect
                        if not d.stad in lStad:
                            lStad.append({'stad': d.stad,'num': qse_count})
                        # Check all the afl for this dialect
                        qsa = Aflevering.objects.filter(entry__in=qse).distinct()
                        for a in qsa:
                            if not a.id in lAfl: lAfl.append(a.id)
                    # Sort the list of afl
                    qsa = Aflevering.objects.filter(id__in=lAfl).select_related().order_by('deel', 'sectie', 'aflnum')
                    lAfl = []
                    # Get a number of standard ranges
                    for oRange in ranges:
                        sBack = get_range(qsa, oRange['d'], oRange['s'])
                        if sBack != "": lAfl.append(sBack)
                    # Get all the afleveringen for this code
                    oDouble = {'code': code, 'count': qs.count(), 'cities': lStad, 'afl_list': lAfl}
                    k_double.append(oDouble)

        context['k_double'] = k_double

        # Return the context
        return context

    def get_queryset(self):
        qs = Dialect.objects.none()
        return qs


class MijnListView(ListView):
    """Listview of mines"""

    model = Mijn
    paginate_by = 10
    template_name = 'dictionary/mijn_list.html'


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(MijnListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET
        search_form = MijnSearchForm(initial)

        context['searchform'] = search_form

        # Determine the count 
        context['entrycount'] = self.get_queryset().count()

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} mijnen".format(THIS_DICTIONARY)

        # Return the calculated context
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in querystring, or use default class property value.
        """
        return self.request.GET.get('paginate_by', self.paginate_by)
        
    def get_queryset(self):

        # Get the parameters passed on with the GET request
        get = self.request.GET.copy()
        get['sortOrder'] = 'naam'

        # Queryset: start out with *ALL* the mines
        qs = Mijn.objects.all()

        # Fine-tuning: search string is the LEMMA
        if 'search' in get and get['search'] != '':
            val = adapt_search(get['search'])
            # The main search is on the NAME of the mine
            query = Q(naam__iregex=val) 

            # check for possible exact numbers having been given
            if re.match('^\d+$', val):
                query = query | Q(sn__exact=val)

            # Apply the filter
            qs = qs.filter(query)

        # Check for toelichting
        if 'toelichting' in get and get['toelichting'] != '':
            val = adapt_search(get['toelichting'])
            # query = Q(nieuw__istartswith=val)
            query = Q(toelichting__iregex=val)
            qs = qs.filter(query)

        # Check for locatie
        if 'locatie' in get and get['locatie'] != '':
            val = adapt_search(get['locatie'])
            # query = Q(nieuw__istartswith=val)
            query = Q(locatie__iregex=val)
            qs = qs.filter(query)

        # Make sure we only have distinct values
        qs = qs.distinct()

        # Sort the queryset by the parameters given
        qs = order_queryset_by_sort_order(get, qs, 'naam')

        # Return the resulting filtered and sorted queryset
        return qs


class DeelListView(ListView):
    """Overview of all the [aflevering] books per [deel]"""

    model = Deel
    template_name = 'dictionary/aflevering_list.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(DeelListView, self).get_context_data(**kwargs)

        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Set the title of the application
        context['title'] = "{} afleveringen".format(THIS_DICTIONARY)

        #context['intro_pdf'] = "wbd-1/1967_Brabantse Dialecten Voorlopige inleiding.pdf"
        #context['intro_op_drie_pdf'] = "wbd-3/2000_Brabantse Dialecten III Inleiding_Compleet.pdf"

        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)

        # Return the calculated context
        return context

    def get_queryset(self):
        qs = Deel.objects.all().order_by('nummer')
        return qs

